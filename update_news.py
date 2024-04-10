from datetime import date
from multiprocessing import Pool

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from tools.spiders import get_news_from_wechat, get_news_from_jiqizhixin
from utils import read_json


def update_news():
    """
    更新新闻存入excel中
    :return:
    """
    wechat_infos = read_json('./configs/wechat_urls.json')
    wechat_urls = []
    for info in wechat_infos:
        wechat_urls.append(info['url'])
    res = []
    with Pool() as pool:
        wechat_res = pool.map_async(get_news_from_wechat, wechat_urls)
        jiqizhixin_res = pool.apply_async(get_news_from_jiqizhixin)
        wechat_news_list = wechat_res.get()
        jiqizhixin_news = jiqizhixin_res.get()
    for wechat_news in wechat_news_list:
        res.extend(wechat_news)
    res.extend(jiqizhixin_news)
    res_df = pd.DataFrame(res, columns=["title", "url", "date"])
    res_df.to_excel(f'./database/db/{str(date.today())}_news.xlsx', index=False)

    # 合并超链接
    # 创建一个新的Excel工作簿
    wb = Workbook()
    ws = wb.active

    # 将DataFrame的数据转换为行，并添加到工作表中
    for r_idx, row in enumerate(dataframe_to_rows(res_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            # 如果是URL列，将相应的标题转换为超链接
            if c_idx == 2:  # 假设URL在第二列
                # 获取相应的标题
                title_cell = ws.cell(row=r_idx, column=c_idx-1)
                # 设置超链接
                ws.cell(row=r_idx, column=c_idx-1).hyperlink = value
                # 为了可视化，通常将超链接的文本设置为蓝色和下划线
                ws.cell(row=r_idx, column=c_idx-1).style = "Hyperlink"
            else:
                ws.cell(row=r_idx, column=c_idx, value=value)

    # 保存工作簿
    file_path = f'./database/links/{str(date.today())}_news_links.xlsx'
    wb.save(file_path)


if __name__ == '__main__':
    update_news()
